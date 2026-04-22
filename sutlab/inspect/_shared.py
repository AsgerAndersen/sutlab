"""
Shared helpers used by multiple inspect modules.
"""

from __future__ import annotations

import dataclasses
from copy import copy
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl.utils
import pandas as pd
from pandas.io.formats.style import Styler

from sutlab.inspect._style import _REL_BASE_SYMBOLS
from sutlab.sut import _match_codes


def _sort_by_id_value(
    df: pd.DataFrame,
    group_levels: list[str],
    sort_id,
) -> pd.DataFrame:
    """Sort non-total rows within groups by sort_id column value, descending.

    Within each group (defined by ``group_levels``), rows where
    ``transaction == ""`` are treated as total/summary rows and kept at the
    end. All other rows are sorted by the value in the ``sort_id`` column,
    largest first. Row order between groups is preserved.

    Parameters
    ----------
    df : pd.DataFrame
        Wide-format inspection table with a MultiIndex containing a
        ``"transaction"`` level and id values as columns.
    group_levels : list of str
        Index level names defining the groups within which rows are sorted.
        Use ``["product"]`` for product detail tables,
        ``["industry"]`` for industry detail tables, and
        ``["product", "price_layer"]`` for price layer tables.
    sort_id : hashable
        Column name (id value, e.g. ``2021``) to sort by.

    Returns
    -------
    pd.DataFrame
        Same DataFrame with rows reordered within each group.
    """
    if df.empty:
        return df

    trans_vals = df.index.get_level_values("transaction")
    total_mask = trans_vals == ""

    level_arrays = {level: df.index.get_level_values(level) for level in group_levels}
    group_tuples = list(dict.fromkeys(
        zip(*[level_arrays[level] for level in group_levels])
    ))

    blocks = []
    for group_key in group_tuples:
        group_mask = level_arrays[group_levels[0]] == group_key[0]
        for level, key in zip(group_levels[1:], group_key[1:]):
            group_mask = group_mask & (level_arrays[level] == key)

        data_rows = df[group_mask & ~total_mask]
        total_rows = df[group_mask & total_mask]

        sorted_data = data_rows.sort_values(by=sort_id, ascending=False)
        blocks.append(pd.concat([sorted_data, total_rows]))

    return pd.concat(blocks)


def _make_sheet_name(field_name: str) -> str:
    """Convert a dataclass field name to an Excel sheet name.

    Excel sheet names are limited to 31 characters. If ``field_name`` is 31
    characters or fewer it is returned unchanged. If it exceeds 31 characters,
    each underscore-separated segment is truncated to its first three
    characters and the segments are rejoined with underscores.

    Parameters
    ----------
    field_name : str
        The dataclass field name to convert.

    Returns
    -------
    str
        A string of at most 31 characters.
    """
    if len(field_name) <= 31:
        return field_name
    segments = field_name.split("_")
    return "_".join(s[:3] for s in segments)


# Field name suffixes that indicate all data columns hold fraction values that
# should be displayed as percentages in Excel (0.05 → 5.0 %).
_PERCENTAGE_FIELD_SUFFIXES = ("_distribution", "_rates", "_growth")

# Default Excel number format strings (1 decimal place).
_EXCEL_NUMBER_FORMAT = "#,##0.0"
_EXCEL_PERCENTAGE_FORMAT = "0.0%"


def _excel_number_format(decimals: int) -> str:
    """Return an Excel number format string for the given decimal count."""
    decimal_part = "0" * decimals
    if decimal_part:
        return f"#,##0.{decimal_part}"
    return "#,##0"


def _excel_percentage_format(decimals: int) -> str:
    """Return an Excel percentage format string for the given decimal count."""
    decimal_part = "0" * decimals
    if decimal_part:
        return f"0.{decimal_part}%"
    return "0%"

# Default width applied to value (non-index) columns.  Wide enough to
# comfortably show a formatted number like "1,234,567.8" (11 chars) plus
# a small margin.
_EXCEL_VALUE_COLUMN_WIDTH = 13


def _apply_number_formats(
    ws,
    df: pd.DataFrame,
    field_name: str,
    display_unit: float | None = None,
    rel_base: int = 100,
    all_rel: bool = False,
    decimals: int = 1,
) -> None:
    """Apply Excel number formats to the data cells of a written worksheet.

    The format applied to each cell depends on the field name and column name:

    - If ``all_rel=True``, all data cells receive the percentage format
      regardless of field name or column name.
    - If ``field_name`` ends with ``_distribution``, ``_rates``, or ``_growth``
      all data cells receive the percentage format (``0.0%``).
    - Otherwise data cells receive the number format (``#,##0.0``), except
      columns whose name starts with ``rel_``, which receive ``0.0%``.

    Non-numeric cells (strings, ``None``) are left unchanged.  Header rows
    and index columns are skipped.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The worksheet to modify, in place.
    df : pd.DataFrame
        The DataFrame that was written to ``ws``.  Used to determine the
        number of header rows (``df.columns.nlevels``) and index columns
        (``df.index.nlevels``), and to identify ``rel_`` column positions.
    field_name : str
        The dataclass field name corresponding to this sheet (e.g.
        ``"balance_distribution"``).  Drives the all-percentage detection.
    all_rel : bool, optional
        When ``True``, all data cells are formatted as percentages regardless
        of ``field_name`` or column names. Default ``False``.
    """
    n_header_rows = df.columns.nlevels
    n_index_cols = df.index.nlevels
    data_cols = list(df.columns)

    is_all_percentage = all_rel or field_name.endswith(_PERCENTAGE_FIELD_SUFFIXES)

    # For "mostly number" tables, track which data-column positions carry
    # relative values and should be formatted as percentages instead.
    rel_col_positions: set[int] = set()
    if not is_all_percentage:
        rel_col_positions = {
            i for i, col in enumerate(data_cols)
            if isinstance(col, str) and col.startswith("rel_")
        }

    for row in ws.iter_rows(
        min_row=n_header_rows + 1,
        min_col=n_index_cols + 1,
    ):
        for col_offset, cell in enumerate(row):
            if not isinstance(cell.value, (int, float)):
                continue
            if is_all_percentage or col_offset in rel_col_positions:
                if rel_base == 100:
                    cell.number_format = _excel_percentage_format(decimals)  # Excel auto-multiplies by 100
                else:
                    symbol = _REL_BASE_SYMBOLS[rel_base]
                    decimal_part = "0" * decimals
                    fmt_str = f"0.{decimal_part}" if decimal_part else "0"
                    cell.value = cell.value * rel_base
                    cell.number_format = f'{fmt_str}"{symbol}"'
            else:
                cell.number_format = _excel_number_format(decimals)
                if display_unit is not None:
                    cell.value = cell.value / display_unit


def _apply_bold_headers(ws, df: pd.DataFrame) -> None:
    """Make all cells in the header rows bold.

    After ``df.to_excel()``, the first ``df.columns.nlevels`` rows hold the
    column headers (including any index-level name cells in those rows).
    This function sets ``bold=True`` on every cell in those rows while
    preserving all other font properties already applied by the Styler.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The worksheet to modify, in place.
    df : pd.DataFrame
        The DataFrame that was written to ``ws``.  Used to determine the
        number of header rows (``df.columns.nlevels``).
    """
    n_header_rows = df.columns.nlevels
    for row in ws.iter_rows(min_row=1, max_row=n_header_rows):
        for cell in row:
            font = copy(cell.font)
            font.bold = True
            cell.font = font


def _set_value_column_widths(ws, n_index_cols: int, n_data_cols: int) -> None:
    """Set a fixed default width on all value (non-index) columns.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The worksheet to modify, in place.
    n_index_cols : int
        Number of leading columns that represent index levels (skipped).
    n_data_cols : int
        Number of data columns to widen.
    """
    for col_idx in range(n_index_cols + 1, n_index_cols + n_data_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = _EXCEL_VALUE_COLUMN_WIDTH


def _fit_index_column_widths(ws, n_index_cols: int) -> None:
    """Set the width of index-level columns to fit their widest cell value.

    Iterates the first ``n_index_cols`` columns in the worksheet, measures
    the string length of every cell value (including the header row), and
    sets the column width to the maximum length plus a small padding of two
    characters.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The worksheet to modify, in place.
    n_index_cols : int
        Number of leading columns that represent index levels.
    """
    for col_idx in range(1, n_index_cols + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2


def _percentile_label(p: float) -> str:
    """Return the canonical display name for a percentile value.

    Parameters
    ----------
    p : float
        Percentile in [0, 1].

    Returns
    -------
    str
        ``"min"`` for 0.0, ``"median"`` for 0.5, ``"max"`` for 1.0,
        ``"p{int(p*100)}"`` for all other values.
    """
    if p == 0.0:
        return "min"
    if p == 0.5:
        return "median"
    if p == 1.0:
        return "max"
    return f"p{int(p * 100)}"


def _build_summary_table(
    detail_table: pd.DataFrame,
    group_levels: list[str],
    item_levels: list[str],
    item_count_label: str,
    total_label: str,
    percentiles: list[float],
    coverage_thresholds: list[float],
) -> pd.DataFrame:
    """Build per-group summary statistics from a detail table.

    Aggregates over the item dimension within each group defined by
    ``group_levels``. Only non-zero item values contribute to counts,
    percentiles, shares, and coverage counts. Non-total rows are identified
    by ``transaction != ""``.

    Parameters
    ----------
    detail_table : pd.DataFrame
        Wide-format DataFrame with a MultiIndex that includes at least a
        ``"transaction"`` level and the levels named in ``group_levels`` and
        ``item_levels``. Non-total rows have ``transaction != ""``.
    group_levels : list of str
        Index level names that define the blocks to aggregate within
        (e.g. ``["industry", "industry_txt", "transaction", "transaction_txt"]``
        for the industries summary, or ``["product", "product_txt"]`` for the
        products summary).
    item_levels : list of str
        Index level names that uniquely identify items within each group
        (e.g. ``["product"]`` for the industries summary, or
        ``["transaction", "category"]`` for the products summary).
    item_count_label : str
        Label for the item count row (e.g. ``"n_products"`` or
        ``"n_categories"``). Coverage rows use this as a prefix:
        ``"{item_count_label}_p{int(t*100)}"``.
    total_label : str
        Label for the total row (e.g. ``"total_supply"`` or ``"total_use"``).
    percentiles : list of float
        Percentile values in [0, 1] to compute for absolute values and shares.
        Value and share rows appear in descending percentile order.
    coverage_thresholds : list of float
        Fraction-of-total thresholds in [0, 1]. For each threshold ``t``,
        the result contains an ``"{item_count_label}_p{int(t*100)}"`` row
        with the minimum number of items (sorted by value descending, per year)
        needed to reach ``t * total``. Coverage rows appear in ascending order.

    Returns
    -------
    pd.DataFrame
        Wide-format summary table with a MultiIndex of
        ``group_levels + ["summary"]``. Columns match ``detail_table.columns``.
        Empty when ``detail_table`` is empty or has no non-total rows.

        Row order within each group block:

        1. ``total_label`` — sum of all items.
        2. ``item_count_label`` — count of non-zero items.
        3. ``"{item_count_label}_p{N}"`` — one row per coverage threshold, ascending.
        4. ``"value_{label}"`` — one row per percentile, descending.
        5. ``"share_{label}"`` — one row per percentile, descending.
    """
    if detail_table.empty:
        return pd.DataFrame()

    id_cols = list(detail_table.columns)

    # Non-total rows: transaction code is non-empty.
    non_total_mask = detail_table.index.get_level_values("transaction") != ""
    item_rows = detail_table[non_total_mask].astype(float)

    if item_rows.empty:
        return pd.DataFrame()

    # --- Vectorised aggregation ---

    # Group totals: sum of all item values per (group, year).
    totals_wide = item_rows.groupby(level=group_levels, dropna=False).sum()

    # n_items: count of non-zero values per (group, year).
    n_items_wide = (item_rows != 0).groupby(level=group_levels, dropna=False).sum()

    # Replace zeros with NaN so groupby quantile skips them.
    nonzero_values = item_rows.where(item_rows != 0)

    # Value percentiles over non-zero items.
    value_quantiles = {
        p: nonzero_values.groupby(level=group_levels, dropna=False).quantile(p)
        for p in percentiles
    }

    # Shares = value / group total. Align group totals to every item row.
    safe_totals = totals_wide.replace(0, float("nan"))
    group_keys = list(
        zip(*[item_rows.index.get_level_values(lv) for lv in group_levels])
    )
    denominators = safe_totals.loc[group_keys].values
    shares = pd.DataFrame(
        item_rows.values / denominators,
        index=item_rows.index,
        columns=id_cols,
    )
    nonzero_shares = shares.where(item_rows != 0)

    # Share percentiles over non-zero items.
    share_quantiles = {
        p: nonzero_shares.groupby(level=group_levels, dropna=False).quantile(p)
        for p in percentiles
    }

    # Coverage counts: for each threshold t, find the minimum number of items
    # (sorted by value descending, per year) whose cumulative sum reaches
    # >= t * total. Computed in a single melt+sort+cumsum pass.
    coverage_wide: dict[float, pd.DataFrame] = {}
    if coverage_thresholds:
        flat = item_rows.reset_index()
        long = flat.melt(
            id_vars=group_levels + item_levels,
            value_vars=id_cols,
            var_name="_year",
            value_name="_value",
        )
        long = long[long["_value"].notna() & (long["_value"] != 0)].copy()

        if not long.empty:
            group_key = group_levels + ["_year"]
            long = long.sort_values(
                group_key + ["_value"],
                ascending=[True] * len(group_key) + [False],
            )
            long["_cumsum"] = long.groupby(group_key, sort=False)["_value"].cumsum()
            long["_total"] = long.groupby(group_key, sort=False)["_value"].transform("sum")
            long["_rank"] = long.groupby(group_key, sort=False).cumcount() + 1

            for t in coverage_thresholds:
                covered = long[long["_cumsum"] >= t * long["_total"]]
                first = (
                    covered.groupby(group_key, sort=False)["_rank"]
                    .first()
                    .reset_index()
                )
                first.columns = group_levels + ["_year", "_count"]
                wide = first.pivot_table(
                    index=group_levels,
                    columns="_year",
                    values="_count",
                    aggfunc="first",
                )
                wide.columns.name = None
                for id_val in id_cols:
                    if id_val not in wide.columns:
                        wide[id_val] = float("nan")
                coverage_wide[t] = wide[id_cols]

    # --- Ordered assembly ---
    n_group = len(group_levels)
    seen: set = set()
    ordered_groups: list = []
    for idx_tuple in item_rows.index:
        key = idx_tuple[:n_group]
        if key not in seen:
            seen.add(key)
            ordered_groups.append(key)

    sorted_thresholds = sorted(coverage_thresholds)
    sorted_percentiles_desc = sorted(percentiles, reverse=True)
    summary_labels = (
        [total_label, item_count_label]
        + [f"{item_count_label}_p{int(t * 100)}" for t in sorted_thresholds]
        + [f"value_{_percentile_label(p)}" for p in sorted_percentiles_desc]
        + [f"share_{_percentile_label(p)}" for p in sorted_percentiles_desc]
    )

    blocks = []
    for group in ordered_groups:
        row_data = [
            totals_wide.loc[group].tolist(),
            n_items_wide.loc[group].tolist(),
        ]
        for t in sorted_thresholds:
            if t in coverage_wide and group in coverage_wide[t].index:
                row_data.append(coverage_wide[t].loc[group].tolist())
            else:
                row_data.append([float("nan")] * len(id_cols))
        for p in sorted_percentiles_desc:
            row_data.append(value_quantiles[p].loc[group].tolist())
        for p in sorted_percentiles_desc:
            row_data.append(share_quantiles[p].loc[group].tolist())

        row_labels = [(*group, label) for label in summary_labels]
        block = pd.DataFrame(
            row_data,
            index=pd.MultiIndex.from_tuples(
                row_labels, names=group_levels + ["summary"]
            ),
            columns=id_cols,
        )
        blocks.append(block)

    if not blocks:
        return pd.DataFrame()

    return pd.concat(blocks)


def _build_growth_table(df: pd.DataFrame) -> pd.DataFrame:
    """Build year-on-year growth table: change relative to the previous year.

    Each value is ``(current - previous) / previous``, so a 5% increase gives
    ``0.05``. The first id column is ``NaN`` throughout. Division by zero also
    yields ``NaN``. Infinite values (from dividing a non-zero change by zero)
    are replaced with ``NaN``.

    Row filtering (e.g. dropping Balance rows) is the caller's responsibility.

    Parameters
    ----------
    df : pd.DataFrame
        Wide-format table with id values as columns.

    Returns
    -------
    pd.DataFrame
        Same shape as ``df`` with growth rates. Empty if ``df`` is empty.
    """
    if df.empty:
        return pd.DataFrame()

    floats = df.astype(float)
    previous = floats.shift(axis=1)
    growth = (floats - previous).div(previous)
    growth = growth.replace([float("inf"), float("-inf")], float("nan"))
    return growth


def _write_inspection_to_excel(inspection_obj: Any, path: str | Path, display_unit: float | None = None, rel_base: int = 100, decimals: int = 1) -> None:
    """Write all non-None tables in an inspection result to an Excel file.

    Each field on ``inspection_obj.data`` that holds a
    :class:`~pandas.DataFrame` is written to a separate sheet. Fields whose
    value is not a DataFrame (e.g. ``None`` or internal non-table attributes)
    are skipped silently. Fields whose value is an empty DataFrame are written
    as empty sheets.

    Where a matching styled property exists on ``inspection_obj``, it is used
    to write the sheet so that colours, number formats, and other Styler
    formatting are preserved in Excel. Otherwise the raw DataFrame is written.

    After writing each sheet, the width of the index-level columns is fitted
    to the widest cell value in that column (including the header row).

    Sheet names are derived from the field name; names exceeding Excel's
    31-character limit are shortened using :func:`_make_sheet_name`.

    Parameters
    ----------
    inspection_obj : inspection result object
        Any inspection result with a ``.data`` attribute that is a dataclass
        (e.g. :class:`~sutlab.inspect.ProductInspection`).
    path : str or Path
        Destination ``.xlsx`` file path.
    """
    path = Path(path)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for f in dataclasses.fields(inspection_obj.data):
            raw = getattr(inspection_obj.data, f.name)
            if not isinstance(raw, pd.DataFrame):
                continue

            sheet_name = _make_sheet_name(f.name)

            try:
                styled = getattr(inspection_obj, f.name, None)
            except Exception:
                styled = None

            if isinstance(styled, Styler):
                try:
                    styled.to_excel(writer, sheet_name=sheet_name)
                except Exception:
                    # Styling failed (e.g. DataFrame structure doesn't match
                    # what the style function expects). Fall back to raw data.
                    raw.to_excel(writer, sheet_name=sheet_name)
            else:
                raw.to_excel(writer, sheet_name=sheet_name)

            ws = writer.sheets[sheet_name]
            all_rel = getattr(inspection_obj, "_all_rel", False)
            _apply_bold_headers(ws, raw)
            _fit_index_column_widths(ws, raw.index.nlevels)
            _set_value_column_widths(ws, raw.index.nlevels, len(raw.columns))
            _apply_number_formats(ws, raw, f.name, display_unit, rel_base, all_rel, decimals)


def _is_protected_row_mask(
    df: pd.DataFrame,
    protected_index_values: dict[str, list],
) -> np.ndarray:
    """Return bool array: True if a row has a protected value in any specified index level."""
    mask = np.zeros(len(df), dtype=bool)
    for level, values in protected_index_values.items():
        if level in df.index.names:
            level_arr = df.index.get_level_values(level)
            for v in values:
                mask |= (level_arr == v)
    return mask


def _iter_group_positions(
    df: pd.DataFrame,
    grouping: list[str] | None,
):
    """Yield positional index arrays for each unique group in ``df``.

    If ``grouping`` is ``None`` or no valid levels are found, yields one
    group covering all rows (global).
    """
    if not grouping:
        yield np.arange(len(df))
        return
    valid = [lv for lv in grouping if lv in df.index.names]
    if not valid:
        yield np.arange(len(df))
        return
    arrays = [df.index.get_level_values(lv) for lv in valid]
    group_tuples = list(dict.fromkeys(zip(*arrays)))
    for group_key in group_tuples:
        group_mask = np.ones(len(df), dtype=bool)
        for arr, key in zip(arrays, group_key):
            group_mask &= (arr == key)
        yield group_mask.nonzero()[0]


def _apply_display_index_filter(
    df: pd.DataFrame,
    display_index: dict[str, list],
    protected_index_values: dict[str, list],
) -> pd.DataFrame:
    """Filter rows by display_index patterns, always keeping protected rows.

    For each level in ``display_index``, keeps only rows whose value at that
    level matches one of the given patterns (using the same pattern syntax as
    :func:`~sutlab.sut.filter_rows`). Rows protected by
    ``protected_index_values`` are always kept regardless of the filter.
    Tables that do not have a given level in their index are left unchanged.
    """
    result = df
    for level, patterns in display_index.items():
        if level not in result.index.names:
            continue
        str_patterns = [str(p) for p in (patterns if isinstance(patterns, list) else [patterns])]
        level_vals_str = result.index.get_level_values(level).astype(str)
        unique_str_vals = list(dict.fromkeys(level_vals_str))
        matched = set(_match_codes(unique_str_vals, str_patterns))
        protected = {str(v) for v in protected_index_values.get(level, [])}
        matched |= protected
        result = result[level_vals_str.isin(matched)]
    return result


def _apply_n_largest_filter(
    df: pd.DataFrame,
    n: int,
    column: str,
    grouping: list[str] | None,
    protected_index_values: dict[str, list],
) -> pd.DataFrame:
    """Keep n largest non-protected rows per group by absolute column value.

    Within each group (defined by ``grouping``), keeps the ``n`` non-protected
    rows with the largest absolute values in ``column``. Protected rows are always kept
    and come after the top-n rows. Rows are returned in their original order
    within each group (sorting is handled separately by ``_apply_column_sort``).
    """
    if df.empty:
        return df
    protected = _is_protected_row_mask(df, protected_index_values)
    keep = np.array(protected)
    col_values = df[column].values
    for positions in _iter_group_positions(df, grouping):
        group_protected = protected[positions]
        non_protected_pos = positions[~group_protected]
        if len(non_protected_pos) == 0:
            continue
        if n >= len(non_protected_pos):
            keep[non_protected_pos] = True
        else:
            group_vals = col_values[non_protected_pos]
            # argsort by absolute value → take last n for largest magnitude
            top_local = np.argsort(np.abs(group_vals))[::-1][:n]
            keep[non_protected_pos[top_local]] = True
    # Preserve original row order
    return df.iloc[keep.nonzero()[0]]


def _apply_column_sort(
    df: pd.DataFrame,
    column: str,
    ascending: bool,
    grouping: list[str] | None,
    protected_index_values: dict[str, list],
) -> pd.DataFrame:
    """Sort non-protected rows within each group by absolute ``column`` value.

    Within each group (defined by ``grouping``), non-protected rows are sorted
    by the absolute value of ``column``; protected rows are appended after them
    in their original relative order.
    """
    if df.empty:
        return df
    protected = _is_protected_row_mask(df, protected_index_values)
    blocks = []
    for positions in _iter_group_positions(df, grouping):
        group_protected = protected[positions]
        data_pos = positions[~group_protected]
        prot_pos = positions[group_protected]
        if len(data_pos) > 0:
            group_df = df.iloc[data_pos].copy()
            group_df["_abs_sort"] = group_df[column].abs()
            sorted_block = group_df.sort_values(by="_abs_sort", ascending=ascending, na_position="last").drop(columns="_abs_sort")
        else:
            sorted_block = df.iloc[[]]
        prot_block = df.iloc[prot_pos] if len(prot_pos) > 0 else df.iloc[[]]
        if len(sorted_block) > 0 or len(prot_block) > 0:
            blocks.append(pd.concat([sorted_block, prot_block]))
    return pd.concat(blocks) if blocks else df.iloc[[]]


def _apply_display_config(
    df: pd.DataFrame,
    table_name: str,
    config,
) -> pd.DataFrame:
    """Apply display configuration to a DataFrame for styled display.

    Returns ``df`` unchanged when the table is in ``protected_tables`` or
    ``df`` is empty. Otherwise applies, in order:

    1. ``display_index`` row filter (pattern-matched, with protected rows always kept).
    2. ``display_values_n_largest`` filter (top-n per group, protected rows always kept).
    3. ``sort_column`` sort (non-protected rows within each group, protected at end).

    Parameters
    ----------
    df : pd.DataFrame
        The raw DataFrame from ``inspection_obj.data``.
    table_name : str
        The field name of this table (used to look up ``protected_tables``
        and ``index_grouping`` in ``config``).
    config : DisplayConfiguration
        The display configuration to apply.
    """
    if df.empty or table_name in config.protected_tables:
        return df
    grouping = config.index_grouping.get(table_name)
    result = df
    if config.display_index:
        result = _apply_display_index_filter(result, config.display_index, config.protected_index_values)
    if config.display_values_n_largest is not None:
        n, column = config.display_values_n_largest
        if column in result.columns:
            result = _apply_n_largest_filter(result, n, column, grouping, config.protected_index_values)
    if config.sort_column is not None and config.sort_column in result.columns:
        result = _apply_column_sort(result, config.sort_column, config.sort_ascending, grouping, config.protected_index_values)
    return result


def _display_index(inspection_obj: Any, values, level: str) -> Any:
    """Filter all tables in ``inspection_obj`` to rows matching ``values`` at ``level``.

    Each DataFrame field whose index contains a level named ``level`` is
    filtered to rows where that level matches one of the given patterns.
    Tables without the named level are left unchanged. ``None`` fields are
    propagated unchanged. Values are converted to strings for matching, so
    integer id values are matched by their string representation.

    Parameters
    ----------
    inspection_obj : inspection result object
        Any inspection result with a ``.data`` attribute that is a dataclass.
    values : str, int, or list
        Values (or patterns) to keep. Accepts the same pattern syntax as
        :func:`~sutlab.sut.filter_rows`: exact, wildcard (``*``), range
        (``:``), negation (``~``). A single value is treated as a one-element
        list.
    level : str
        Name of the index level to filter on.

    Returns
    -------
    Same type as ``inspection_obj``
        A new inspection result of the same class with filtered tables.
    """
    patterns = [values] if not isinstance(values, list) else list(values)
    str_patterns = [str(p) for p in patterns]

    filtered_fields: dict = {}
    for f in dataclasses.fields(inspection_obj.data):
        val = getattr(inspection_obj.data, f.name)
        if val is None or not isinstance(val, pd.DataFrame):
            filtered_fields[f.name] = val
            continue
        if val.empty or level not in val.index.names:
            filtered_fields[f.name] = val
            continue

        level_vals_str = val.index.get_level_values(level).astype(str)
        unique_str_vals = list(dict.fromkeys(level_vals_str))
        matched = set(_match_codes(unique_str_vals, str_patterns))
        filtered_fields[f.name] = val[level_vals_str.isin(matched)]

    data_cls = type(inspection_obj.data)
    new_data = data_cls(**filtered_fields)
    return dataclasses.replace(inspection_obj, data=new_data)


def _get_index_values(
    inspection_obj: Any,
    table: str,
    levels: str | list[str],
    *,
    as_list: bool = False,
) -> pd.DataFrame | list:
    """Return unique index value combinations for a table after applying display config.

    Parameters
    ----------
    inspection_obj : inspection result object
        Any inspection result with a ``.data`` attribute that is a dataclass
        and a ``.display_configuration`` attribute.
    table : str
        Name of a DataFrame field on ``inspection_obj.data``.
    levels : str or list of str
        One or more index level names whose unique combinations to return.
    as_list : bool, default False
        If ``True``, return a plain list of unique values. Requires exactly
        one level; raises ``ValueError`` if more than one level is requested.

    Returns
    -------
    pd.DataFrame or list
        When ``as_list=False``: one column per requested level, unique
        combinations only. Rows where all values are ``""`` or ``NaN`` are
        dropped. Index is a default RangeIndex.
        When ``as_list=True``: a plain list of unique values for the single
        requested level.

    Raises
    ------
    ValueError
        If ``table`` is not a DataFrame field on ``inspection_obj.data``,
        if the field is ``None``, if any of ``levels`` is not an index
        level of that table, or if ``as_list=True`` and more than one level
        is requested.
    """
    data_field_names = {f.name for f in dataclasses.fields(inspection_obj.data)}
    if table not in data_field_names:
        available = sorted(data_field_names)
        raise ValueError(
            f"Table {table!r} not found. Available tables: {available}."
        )

    df = getattr(inspection_obj.data, table)

    if df is None:
        raise ValueError(
            f"Table {table!r} is None (no data available for this table)."
        )

    if not isinstance(df, pd.DataFrame):
        raise ValueError(
            f"Table {table!r} is not a DataFrame."
        )

    df = _apply_display_config(df, table, inspection_obj.display_configuration)

    level_list = [levels] if isinstance(levels, str) else list(levels)

    if as_list and len(level_list) > 1:
        raise ValueError(
            f"as_list=True requires exactly one level, but {len(level_list)} were given: {level_list}."
        )

    missing = [lv for lv in level_list if lv not in df.index.names]
    if missing:
        raise ValueError(
            f"Level(s) {missing} not found in table {table!r}. "
            f"Available levels: {list(df.index.names)}."
        )

    result = df.index.to_frame(index=False)[level_list].drop_duplicates().reset_index(drop=True)

    all_empty = result.apply(lambda col: col.isna() | (col == ""), axis=0).all(axis=1)
    result = result[~all_empty].reset_index(drop=True)

    if as_list:
        return result[level_list[0]].tolist()
    return result
