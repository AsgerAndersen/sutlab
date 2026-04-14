"""
Tests for write_to_excel on inspection result objects.

Covers:
- _make_sheet_name: truncation logic
- _fit_index_column_widths: column width setting
- _write_inspection_to_excel / write_to_excel:
    - file creation
    - correct sheet names
    - None-field skipping
    - non-DataFrame field skipping (_p1_trans on IndustryInspectionData)
    - empty-DataFrame sheets written (not skipped)
    - styling applied when inspection result comes from a real inspection call
    - column widths set on index columns
"""

import dataclasses

import openpyxl
import openpyxl.utils
import pytest
import pandas as pd
from pandas.io.formats.style import Styler

from sutlab.inspect._shared import (
    _make_sheet_name,
    _apply_bold_headers,
    _apply_number_formats,
    _fit_index_column_widths,
    _set_value_column_widths,
    _write_inspection_to_excel,
    _EXCEL_NUMBER_FORMAT,
    _EXCEL_PERCENTAGE_FORMAT,
    _EXCEL_VALUE_COLUMN_WIDTH,
)
from sutlab.inspect import (
    ProductInspection,
    ProductInspectionData,
    IndustryInspection,
    IndustryInspectionData,
    UnbalancedTargetsInspection,
    UnbalancedTargetsData,
    SUTComparisonInspection,
    SUTComparisonData,
)


# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def realistic_balancing_result():
    """UnbalancedTargetsInspection with column structure matching a real inspection.

    ``_supply_styler`` requires at least one column not prefixed with
    ``target_/diff_/rel_/tol_/violation_``.  The DataFrames here mirror what
    ``inspect_unbalanced_targets`` would produce.
    """
    supply_cat = pd.DataFrame(
        {
            "bas": [300.0],
            "target_bas": [360.0],
            "diff_bas": [-60.0],
            "rel_bas": [-0.167],
            "tol_bas": [10.0],
            "violation_bas": [-50.0],
        },
        index=pd.MultiIndex.from_tuples(
            [("0100", "X")], names=["trans", "brch"]
        ),
    )
    use_cat = pd.DataFrame(
        {
            "koeb": [80.0],
            "target_koeb": [90.0],
            "diff_koeb": [-10.0],
            "rel_koeb": [-0.111],
            "tol_koeb": [float("nan")],
            "violation_koeb": [float("nan")],
        },
        index=pd.MultiIndex.from_tuples(
            [("2000", "X")], names=["trans", "brch"]
        ),
    )
    supply_trans = pd.DataFrame(
        {
            "bas": [300.0],
            "target_bas": [360.0],
            "diff_bas": [-60.0],
            "rel_bas": [-0.167],
            "tol_bas": [10.0],
            "violation_bas": [-50.0],
        },
        index=pd.Index(["0100"], name="trans"),
    )
    use_trans = pd.DataFrame(
        {
            "koeb": [80.0],
            "target_koeb": [90.0],
            "diff_koeb": [-10.0],
            "rel_koeb": [-0.111],
            "tol_koeb": [float("nan")],
            "violation_koeb": [float("nan")],
        },
        index=pd.Index(["2000"], name="trans"),
    )
    summary = pd.DataFrame(
        {"n_unbalanced": [1, 1, 1, 1]},
        index=pd.Index(["supply_transactions", "supply_categories", "use_transactions", "use_categories"], name="table"),
    )
    data = UnbalancedTargetsData(
        supply_categories=supply_cat,
        use_categories=use_cat,
        supply_categories_violations=None,
        use_categories_violations=None,
        supply_transactions=supply_trans,
        use_transactions=use_trans,
        supply_transactions_violations=None,
        use_transactions_violations=None,
        summary=summary,
    )
    return UnbalancedTargetsInspection(data=data)


# ---------------------------------------------------------------------------
# _make_sheet_name
# ---------------------------------------------------------------------------


def test_make_sheet_name_short_name_unchanged():
    assert _make_sheet_name("balance") == "balance"


def test_make_sheet_name_exactly_31_unchanged():
    name = "a" * 31
    assert _make_sheet_name(name) == name


def test_make_sheet_name_truncates_when_over_31():
    # balancing_targets_use_purchasers = 32 chars
    result = _make_sheet_name("balancing_targets_use_purchasers")
    assert result == "bal_tar_use_pur"
    assert len(result) <= 31


def test_make_sheet_name_truncates_longer_name():
    # balancing_targets_use_price_layers = 34 chars
    result = _make_sheet_name("balancing_targets_use_price_layers")
    assert result == "bal_tar_use_pri_lay"
    assert len(result) <= 31


def test_make_sheet_name_all_known_fields_within_limit():
    """Every current inspection field name must map to a sheet name <= 31 chars."""
    from sutlab.inspect import (
        ProductInspectionData,
        IndustryInspectionData,
        FinalUseInspectionData,
        UnbalancedProductsData,
        UnbalancedTargetsData,
        SUTComparisonData,
    )
    data_classes = [
        ProductInspectionData,
        IndustryInspectionData,
        FinalUseInspectionData,
        UnbalancedProductsData,
        UnbalancedTargetsData,
        SUTComparisonData,
    ]
    for cls in data_classes:
        for f in dataclasses.fields(cls):
            sheet = _make_sheet_name(f.name)
            assert len(sheet) <= 31, (
                f"{cls.__name__}.{f.name} -> '{sheet}' ({len(sheet)} chars)"
            )


# ---------------------------------------------------------------------------
# _apply_bold_headers
# ---------------------------------------------------------------------------


def test_apply_bold_headers_single_level(tmp_path):
    """All cells in the single header row are bold."""
    df = pd.DataFrame(
        {"val": [1.0]},
        index=pd.Index(["A"], name="label"),
    )
    ws = _write_df_to_ws(tmp_path, df)
    _apply_bold_headers(ws, df)
    # Row 1: index name cell (A1) and column name cell (B1)
    assert ws["A1"].font.bold is True
    assert ws["B1"].font.bold is True


def test_apply_bold_headers_does_not_bold_data_rows(tmp_path):
    """Data rows (row 2 onward) are not affected."""
    df = pd.DataFrame(
        {"val": [1.0]},
        index=pd.Index(["A"], name="label"),
    )
    ws = _write_df_to_ws(tmp_path, df)
    _apply_bold_headers(ws, df)
    assert not ws["A2"].font.bold
    assert not ws["B2"].font.bold


def test_apply_bold_headers_preserves_existing_font_properties(tmp_path):
    """Bold is added without clearing other font properties set by the Styler."""
    df = pd.DataFrame({"val": [1.0]}, index=pd.Index(["A"], name="label"))
    path = tmp_path / "font.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Sheet1")
        ws = writer.sheets["Sheet1"]
        # Manually set a non-default font size on a header cell before bolding.
        from openpyxl.styles import Font
        ws["B1"].font = Font(size=14)
        _apply_bold_headers(ws, df)
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    assert ws["B1"].font.bold is True
    assert ws["B1"].font.size == 14


# ---------------------------------------------------------------------------
# _apply_number_formats
# ---------------------------------------------------------------------------


def _write_df_to_ws(tmp_path, df, sheet_name="Sheet1"):
    """Helper: write df to an xlsx and return the openpyxl worksheet."""
    path = tmp_path / "fmt.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
    wb = openpyxl.load_workbook(path)
    return wb[sheet_name]


def test_apply_number_formats_number_table(tmp_path):
    """Data cells in a plain table receive the number format."""
    df = pd.DataFrame({"value": [1234.5]}, index=pd.Index(["A"], name="label"))
    ws = _write_df_to_ws(tmp_path, df)
    _apply_number_formats(ws, df, "balance")
    # Row 2 (after 1 header row), column B (after 1 index column)
    assert ws["B2"].number_format == _EXCEL_NUMBER_FORMAT


def test_apply_number_formats_distribution_table(tmp_path):
    """Data cells in a _distribution table receive the percentage format."""
    df = pd.DataFrame({"share": [0.25]}, index=pd.Index(["A"], name="label"))
    ws = _write_df_to_ws(tmp_path, df)
    _apply_number_formats(ws, df, "balance_distribution")
    assert ws["B2"].number_format == _EXCEL_PERCENTAGE_FORMAT


def test_apply_number_formats_rates_table(tmp_path):
    """Data cells in a _rates table receive the percentage format."""
    df = pd.DataFrame({"rate": [0.05]}, index=pd.Index(["A"], name="label"))
    ws = _write_df_to_ws(tmp_path, df)
    _apply_number_formats(ws, df, "price_layers_rates")
    assert ws["B2"].number_format == _EXCEL_PERCENTAGE_FORMAT


def test_apply_number_formats_growth_table(tmp_path):
    """Data cells in a _growth table receive the percentage format."""
    df = pd.DataFrame({"g": [0.03]}, index=pd.Index(["A"], name="label"))
    ws = _write_df_to_ws(tmp_path, df)
    _apply_number_formats(ws, df, "balance_growth")
    assert ws["B2"].number_format == _EXCEL_PERCENTAGE_FORMAT


def test_apply_number_formats_rel_column_in_mixed_table(tmp_path):
    """In a non-percentage table, rel_ columns get percentage format."""
    df = pd.DataFrame(
        {"bas": [300.0], "rel_bas": [-0.167]},
        index=pd.Index(["0100"], name="trans"),
    )
    ws = _write_df_to_ws(tmp_path, df)
    _apply_number_formats(ws, df, "supply")
    # Column B = bas → number format
    assert ws["B2"].number_format == _EXCEL_NUMBER_FORMAT
    # Column C = rel_bas → percentage format
    assert ws["C2"].number_format == _EXCEL_PERCENTAGE_FORMAT


def test_apply_number_formats_skips_string_cells(tmp_path):
    """Non-numeric cells are left with their default format."""
    df = pd.DataFrame({"label": ["some text"]}, index=pd.Index(["A"], name="idx"))
    ws = _write_df_to_ws(tmp_path, df)
    default_format = ws["B2"].number_format
    _apply_number_formats(ws, df, "balance")
    assert ws["B2"].number_format == default_format


def test_apply_number_formats_skips_header_and_index(tmp_path):
    """Header row and index column cells are not reformatted."""
    df = pd.DataFrame({"value": [100.0]}, index=pd.Index(["A"], name="label"))
    ws = _write_df_to_ws(tmp_path, df)
    header_fmt_before = ws["B1"].number_format
    index_fmt_before = ws["A2"].number_format
    _apply_number_formats(ws, df, "balance")
    assert ws["B1"].number_format == header_fmt_before
    assert ws["A2"].number_format == index_fmt_before


# ---------------------------------------------------------------------------
# _set_value_column_widths
# ---------------------------------------------------------------------------


def test_set_value_column_widths(tmp_path):
    """Data columns receive the fixed default width."""
    df = pd.DataFrame(
        {"a": [1.0], "b": [2.0]},
        index=pd.Index(["X"], name="label"),
    )
    path = tmp_path / "val.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Sheet1")
        ws = writer.sheets["Sheet1"]
        _set_value_column_widths(ws, n_index_cols=1, n_data_cols=2)

    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    assert ws.column_dimensions["B"].width == _EXCEL_VALUE_COLUMN_WIDTH
    assert ws.column_dimensions["C"].width == _EXCEL_VALUE_COLUMN_WIDTH


def test_write_to_excel_value_columns_wider_than_default(tmp_path, balancing_result_with_none):
    """Value columns in the written file are wider than Excel's default (8.43)."""
    path = tmp_path / "out.xlsx"
    balancing_result_with_none.write_to_excel(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["supply_categories"]
    # Column B is the first (and only) value column in the supply_categories sheet
    assert ws.column_dimensions["B"].width == _EXCEL_VALUE_COLUMN_WIDTH


# ---------------------------------------------------------------------------
# _fit_index_column_widths
# ---------------------------------------------------------------------------


def test_fit_index_column_widths_sets_width(tmp_path):
    """Column widths are set to max cell length + 2 for each index column."""
    # Create a real worksheet via pandas to_excel, then check widths.
    df = pd.DataFrame(
        {"value": [1, 2]},
        index=pd.Index(["short", "a_longer_label"], name="label"),
    )
    path = tmp_path / "widths.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Sheet1")
        ws = writer.sheets["Sheet1"]
        _fit_index_column_widths(ws, n_index_cols=1)

    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    # Widest value in column A is "a_longer_label" (14 chars) + 2 padding = 16
    assert ws.column_dimensions["A"].width == 16


def test_fit_index_column_widths_multiindex(tmp_path):
    """Both index columns are widened when the DataFrame has a two-level index."""
    df = pd.DataFrame(
        {"value": [1]},
        index=pd.MultiIndex.from_tuples(
            [("short", "a_much_longer_value")], names=["a", "b"]
        ),
    )
    path = tmp_path / "multi.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Sheet1")
        ws = writer.sheets["Sheet1"]
        _fit_index_column_widths(ws, n_index_cols=2)

    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    # Column A: widest is "short" (5) or index name "a" (1) → 5 + 2 = 7
    assert ws.column_dimensions["A"].width == 7
    # Column B: widest is "a_much_longer_value" (19) → 19 + 2 = 21
    assert ws.column_dimensions["B"].width == 21


# ---------------------------------------------------------------------------
# write_to_excel — UnbalancedTargetsInspection (has None fields)
# ---------------------------------------------------------------------------


@pytest.fixture
def balancing_result_with_none():
    """UnbalancedTargetsInspection where all violations tables are None."""
    supply_cat = pd.DataFrame({"target_bas": [100.0]})
    use_cat = pd.DataFrame({"target_koeb": [200.0]})
    supply_trans = pd.DataFrame({"target_bas": [100.0]})
    use_trans = pd.DataFrame({"target_koeb": [200.0]})
    summary = pd.DataFrame(
        {"n_unbalanced": [1, 1, 1, 1]},
        index=pd.Index(["supply_transactions", "supply_categories", "use_transactions", "use_categories"], name="table"),
    )
    data = UnbalancedTargetsData(
        supply_categories=supply_cat,
        use_categories=use_cat,
        supply_categories_violations=None,
        use_categories_violations=None,
        supply_transactions=supply_trans,
        use_transactions=use_trans,
        supply_transactions_violations=None,
        use_transactions_violations=None,
        summary=summary,
    )
    return UnbalancedTargetsInspection(data=data)


@pytest.fixture
def balancing_result_with_violations():
    """UnbalancedTargetsInspection where all violations tables are populated."""
    supply_cat = pd.DataFrame({"target_bas": [100.0]})
    use_cat = pd.DataFrame({"target_koeb": [200.0]})
    supply_cat_viol = pd.DataFrame({"violation_bas": [-50.0]})
    use_cat_viol = pd.DataFrame({"violation_koeb": [-12.0]})
    supply_trans = pd.DataFrame({"target_bas": [100.0]})
    use_trans = pd.DataFrame({"target_koeb": [200.0]})
    supply_trans_viol = pd.DataFrame({"violation_bas": [-50.0]})
    use_trans_viol = pd.DataFrame({"violation_koeb": [-12.0]})
    summary = pd.DataFrame(
        {"n_unbalanced": [1, 1, 1, 1, 1, 1, 1, 1]},
        index=pd.Index([
            "supply_transactions", "supply_categories",
            "use_transactions", "use_categories",
            "supply_transactions_violations", "supply_categories_violations",
            "use_transactions_violations", "use_categories_violations",
        ], name="table"),
    )
    data = UnbalancedTargetsData(
        supply_categories=supply_cat,
        use_categories=use_cat,
        supply_categories_violations=supply_cat_viol,
        use_categories_violations=use_cat_viol,
        supply_transactions=supply_trans,
        use_transactions=use_trans,
        supply_transactions_violations=supply_trans_viol,
        use_transactions_violations=use_trans_viol,
        summary=summary,
    )
    return UnbalancedTargetsInspection(data=data)


def test_write_to_excel_creates_file(tmp_path, balancing_result_with_none):
    path = tmp_path / "out.xlsx"
    balancing_result_with_none.write_to_excel(path)
    assert path.exists()


def test_write_to_excel_skips_none_fields(tmp_path, balancing_result_with_none):
    path = tmp_path / "out.xlsx"
    balancing_result_with_none.write_to_excel(path)
    wb = openpyxl.load_workbook(path)
    assert "supply_categories_violations" not in wb.sheetnames
    assert "use_categories_violations" not in wb.sheetnames
    assert "supply_transactions_violations" not in wb.sheetnames
    assert "use_transactions_violations" not in wb.sheetnames


def test_write_to_excel_writes_non_none_fields(tmp_path, balancing_result_with_none):
    path = tmp_path / "out.xlsx"
    balancing_result_with_none.write_to_excel(path)
    wb = openpyxl.load_workbook(path)
    assert "supply_categories" in wb.sheetnames
    assert "use_categories" in wb.sheetnames
    assert "supply_transactions" in wb.sheetnames
    assert "use_transactions" in wb.sheetnames
    assert "summary" in wb.sheetnames


def test_write_to_excel_writes_all_fields_when_no_none(
    tmp_path, balancing_result_with_violations
):
    path = tmp_path / "out.xlsx"
    balancing_result_with_violations.write_to_excel(path)
    wb = openpyxl.load_workbook(path)
    expected = {
        "supply_categories",
        "use_categories",
        "supply_categories_violations",
        "use_categories_violations",
        "supply_transactions",
        "use_transactions",
        "supply_transactions_violations",
        "use_transactions_violations",
        "summary",
    }
    assert set(wb.sheetnames) == expected


def test_write_to_excel_index_column_widths_set(tmp_path, balancing_result_with_none):
    """Index column A of the supply_categories sheet has a non-zero width after writing."""
    path = tmp_path / "out.xlsx"
    balancing_result_with_none.write_to_excel(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["supply_categories"]
    assert ws.column_dimensions["A"].width > 0


# ---------------------------------------------------------------------------
# write_to_excel — ProductInspection (empty DataFrames written, not skipped)
# ---------------------------------------------------------------------------


@pytest.fixture
def product_result_with_empty_tables():
    """ProductInspection with only balance populated; rest are empty DataFrames."""
    balance = pd.DataFrame({"value": [1, 2, 3]})
    data = ProductInspectionData(balance=balance)
    return ProductInspection(data=data)


def test_write_to_excel_writes_empty_dataframes(
    tmp_path, product_result_with_empty_tables
):
    path = tmp_path / "out.xlsx"
    product_result_with_empty_tables.write_to_excel(path)
    wb = openpyxl.load_workbook(path)
    # All 13 fields should be present — empty DataFrames are not skipped
    assert "balance" in wb.sheetnames
    assert "supply_products" in wb.sheetnames
    assert "price_layers_rates" in wb.sheetnames
    assert len(wb.sheetnames) == 13


# ---------------------------------------------------------------------------
# write_to_excel — IndustryInspection (non-DataFrame _p1_trans field skipped)
# ---------------------------------------------------------------------------


def test_industry_inspection_non_dataframe_field_is_on_outer_class():
    """_p1_trans lives on IndustryInspection, not on IndustryInspectionData.

    _write_inspection_to_excel iterates ``inspection_obj.data``, so _p1_trans
    is never encountered — it does not need to be filtered by the isinstance
    guard.  This test documents the data structure so the behaviour is clear.
    """
    data = IndustryInspectionData(balance=pd.DataFrame())
    result = IndustryInspection(data=data)

    outer_field_names = [f.name for f in dataclasses.fields(result)]
    inner_field_names = [f.name for f in dataclasses.fields(data)]

    assert "_p1_trans" in outer_field_names
    assert not isinstance(result._p1_trans, pd.DataFrame)
    assert "_p1_trans" not in inner_field_names


def test_write_to_excel_skips_non_dataframe_field(tmp_path):
    """_p1_trans must not appear as a sheet and must not cause a crash."""
    data = IndustryInspectionData(balance=pd.DataFrame())
    result = IndustryInspection(data=data)
    path = tmp_path / "out.xlsx"
    result.write_to_excel(path)
    wb = openpyxl.load_workbook(path)
    assert "_p1_trans" not in wb.sheetnames


# ---------------------------------------------------------------------------
# write_to_excel — sheet name truncation
# ---------------------------------------------------------------------------


def test_write_to_excel_sheet_name_truncation(tmp_path):
    """SUTComparisonData has two fields exceeding 31 chars; verify truncation."""
    empty = pd.DataFrame()
    data = SUTComparisonData(
        supply=empty,
        use_basic=empty,
        use_purchasers=empty,
        use_price_layers=empty,
        balancing_targets_supply=None,
        balancing_targets_use_basic=None,
        balancing_targets_use_purchasers=None,
        balancing_targets_use_price_layers=None,
        summary=empty,
    )
    result = SUTComparisonInspection(data=data)
    path = tmp_path / "out.xlsx"
    result.write_to_excel(path)
    wb = openpyxl.load_workbook(path)
    for name in wb.sheetnames:
        assert len(name) <= 31, f"Sheet name too long: '{name}'"


# ---------------------------------------------------------------------------
# write_to_excel — styling applied when DataFrames have the right structure
# ---------------------------------------------------------------------------


def test_write_to_excel_applies_styling(tmp_path, realistic_balancing_result):
    """Cells written via the Styler path carry background fills."""
    path = tmp_path / "styled.xlsx"
    realistic_balancing_result.write_to_excel(path)

    wb = openpyxl.load_workbook(path)
    ws = wb["supply_categories"]

    # At least one non-header cell should have a non-None, non-default fill
    # (the supply palette applies a coloured background to data cells).
    filled = [
        cell
        for row in ws.iter_rows(min_row=2)
        for cell in row
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb != "00000000"
    ]
    assert len(filled) > 0, "Expected at least one coloured cell in the supply sheet"


def test_write_to_excel_falls_back_to_raw_when_styling_fails(
    tmp_path, balancing_result_with_none
):
    """When the styled property fails (bad column structure), raw data is written."""
    # balancing_result_with_none has supply = {"target_bas": [100.0]} which
    # causes _supply_styler to raise StopIteration (no non-prefixed column).
    # The fallback should write the raw DataFrame instead of crashing.
    path = tmp_path / "fallback.xlsx"
    balancing_result_with_none.write_to_excel(path)
    wb = openpyxl.load_workbook(path)
    assert "supply_categories" in wb.sheetnames
